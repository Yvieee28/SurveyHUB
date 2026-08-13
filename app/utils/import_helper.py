import pandas as pd
from io import BytesIO
from flask_babel import _
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from app import db
from app.models.response import Response, Answer, Company
from app.models.survey import Survey, Block, Question, Choice
from app.utils.validation import validate_email, validate_siret, validate_phone


def is_survey_importable(survey):
    """Check if survey is published and all questions are multiple_choice."""
    if survey.status != 'published':
        return False, _("Survey must be published to import responses.")
    
    for block in survey.blocks:
        for question in block.questions:
            if question.question_type != 'multiple_choice':
                return False, _("Only surveys with multiple-choice questions can be imported.")
    
    return True, ""


def get_survey_questions(survey):
    """Returns ordered list of questions across all blocks."""
    questions = []
    for block in sorted(survey.blocks, key=lambda b: b.position):
        for question in sorted(block.questions, key=lambda q: q.position):
            questions.append(question)
    return questions


def generate_import_template(survey):
    """Build Excel template with OpenPyXL, returns (BytesIO, filename)."""
    questions = get_survey_questions(survey)
    
    # Build headers
    headers = [_('Company Name'), _('SIRET'), _('Email'), _('Phone')]
    for q in questions:
        headers.append(q.text)
    
    # Create workbook
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = _("Import Template")
    
    # Write headers with formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add comments to question columns with valid choices
    for i, q in enumerate(questions):
        col = i + 5  # Columns E onwards (1-indexed, 4 fixed columns)
        choices = [c.choice_text for c in sorted(q.choices, key=lambda c: c.position)]
        comment_text = _("Valid choices:") + "\n" + "\n".join(f"• {c}" for c in choices)
        comment = Comment(comment_text, "CCI SurveyHUB")
        comment.width = 300
        comment.height = 150
        ws.cell(row=1, column=col).comment = comment
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    for i in range(len(questions)):
        col_letter = chr(69 + i) if i < 23 else None  # E onwards
        if col_letter:
            ws.column_dimensions[col_letter].width = 25
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{survey.title.replace(' ', '_')}_import_template.xlsx"
    return output, filename


def validate_import_file(survey, file_storage):
    """Read Excel file, validate headers and rows. Returns dict with valid/rejected lists."""
    questions = get_survey_questions(survey)
    
    # Expected headers
    expected_headers = [_('Company Name'), _('SIRET'), _('Email'), _('Phone')]
    for q in questions:
        expected_headers.append(q.text)
    
    try:
        df = pd.read_excel(file_storage, engine='openpyxl')
    except Exception as e:
        return {'error': str(_('Invalid Excel file: %(error)s', error=str(e)))}
    
    # Check headers
    actual_headers = [str(h).strip() for h in df.columns.tolist()]
    expected_headers_stripped = [h.strip() for h in expected_headers]
    
    if actual_headers != expected_headers_stripped:
        return {'error': _('Excel headers do not match the survey structure. Please download the current template.')}
    
    # Build choice lookup: question_text -> set of valid choice texts
    choice_lookup = {}
    for q in questions:
        choice_lookup[q.text] = {c.choice_text for c in q.choices}
    
    # Build question lookup by text
    question_lookup = {q.text: q for q in questions}
    
    # Get existing SIRETs that already have responses for this survey
    existing_responses = Response.query.filter_by(survey_id=survey.id).all()
    existing_company_ids = {r.company_id for r in existing_responses}
    existing_sirets = set()
    if existing_company_ids:
        companies = Company.query.filter(Company.id.in_(existing_company_ids)).all()
        existing_sirets = {c.siret for c in companies if c.siret}
    
    valid_rows = []
    rejected_rows = []
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed, header is row 1)
        company_name = str(row.get(_('Company Name'), '')).strip() if pd.notna(row.get(_('Company Name'))) else ''
        siret = str(row.get(_('SIRET'), '')).strip() if pd.notna(row.get(_('SIRET'))) else ''
        email = str(row.get(_('Email'), '')).strip() if pd.notna(row.get(_('Email'))) else ''
        phone = str(row.get(_('Phone'), '')).strip() if pd.notna(row.get(_('Phone'))) else ''
        
        # Validate required fields
        errors = []
        
        if not company_name or company_name == 'nan':
            errors.append(_("Company name is required"))
        
        if not siret or siret == 'nan':
            errors.append(_("SIRET is required"))
        else:
            # Normalize SIRET
            siret_digits = ''.join(filter(str.isdigit, siret))
            if len(siret_digits) != 14:
                errors.append(_("SIRET must be exactly 14 digits"))
            siret = siret_digits
        
        # Validate email if present
        if email and email != 'nan':
            email_valid, email_err = validate_email(email)
            if not email_valid:
                errors.append(email_err)
        else:
            email = ''
        
        # Validate phone if present
        if phone and phone != 'nan':
            phone_valid, phone_err = validate_phone(phone)
            if not phone_valid:
                errors.append(phone_err)
            else:
                # Normalize phone
                phone = ''.join(filter(str.isdigit, phone))
        else:
            phone = ''
        
        # Validate question answers
        answers = {}
        for q in questions:
            answer_val = str(row.get(q.text, '')).strip() if pd.notna(row.get(q.text)) else ''
            if answer_val and answer_val != 'nan':
                if answer_val not in choice_lookup[q.text]:
                    errors.append(_("Invalid choice '%(answer)s' for question: %(question)s", answer=answer_val, question=q.text[:50]))
                else:
                    # Find the choice_id
                    choice = next((c for c in q.choices if c.choice_text == answer_val), None)
                    if choice:
                        answers[q.id] = choice.id
        
        # Check for duplicate SIRET response
        if siret and siret in existing_sirets:
            rejected_rows.append({
                'row': row_num,
                'company_name': company_name,
                'siret': siret,
                'status': _('Duplicate'),
                'reason': _('SIRET already has a response for this survey')
            })
            continue
        
        if errors:
            rejected_rows.append({
                'row': row_num,
                'company_name': company_name,
                'siret': siret,
                'status': _('Invalid'),
                'reason': '; '.join(errors)
            })
        else:
            valid_rows.append({
                'row': row_num,
                'company_name': company_name,
                'siret': siret,
                'email': email,
                'phone': phone,
                'answers': answers
            })
            # Add to existing_sirets to catch duplicates within the file
            if siret:
                existing_sirets.add(siret)
    
    return {
        'valid': valid_rows,
        'rejected': rejected_rows,
        'total': len(df)
    }


def commit_import(survey, valid_rows):
    """Create Company/Response/Answer records for valid rows. Returns count of imported."""
    imported_count = 0
    questions = get_survey_questions(survey)
    required_q_ids = {q.id for q in questions if q.required}
    
    for row in valid_rows:
        # Find or create company
        company = Company.query.filter_by(siret=row['siret']).first()
        if not company:
            company = Company(
                company_name=row['company_name'],
                siret=row['siret'],
                email=row['email'],
                phone=row['phone']
            )
            db.session.add(company)
            db.session.flush()
        else:
            # Update company info
            company.company_name = row['company_name']
            if row['email']:
                company.email = row['email']
            if row['phone']:
                company.phone = row['phone']
        
        # Determine completion status
        answered_q_ids = set(row['answers'].keys())
        is_complete = required_q_ids.issubset(answered_q_ids)
        
        # Create response
        response = Response(
            survey_id=survey.id,
            company_id=company.id,
            completion_status='complete' if is_complete else 'incomplete'
        )
        db.session.add(response)
        db.session.flush()
        
        # Create answers
        for question_id, choice_id in row['answers'].items():
            answer = Answer(
                response_id=response.id,
                question_id=question_id,
                choice_id=choice_id
            )
            db.session.add(answer)
        
        imported_count += 1
    
    db.session.commit()
    return imported_count


def generate_rejected_report(rejected_rows):
    """Create downloadable CSV of rejected rows. Returns (BytesIO, filename)."""
    if not rejected_rows:
        return None, None
    
    df = pd.DataFrame(rejected_rows)
    output = BytesIO()
    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    
    filename = "rejected_rows.csv"
    return output, filename
